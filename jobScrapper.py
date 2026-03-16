import argparse
import logging
import random
import re
import time
from datetime import datetime
from urllib.parse import quote

import pandas as pd
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

# Pour les couleurs Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("ERREUR: openpyxl non installé - impossible de générer le fichier Excel")
    print("Installez-le avec: pip install openpyxl")


# Mapping durée → paramètre LinkedIn f_TPR
DURATION_MAP = {
    "1d": "r86400",    # 1 jour
    "1w": "r604800",   # 1 semaine
}


class LinkedInJobScraper:
    def __init__(self, use_selenium=False, duration="1w"):
        self.use_selenium = use_selenium
        self.duration = duration
        self.session = requests.Session()
        self.jobs_data = []
        
        # Headers pour imiter un navigateur réel
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'fr-FR,fr;q=0.5',
            'Accept-Encoding': 'gzip, deflate',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
        }
        
        if self.use_selenium:
            self.setup_selenium()
    
    def setup_selenium(self):
        """Configure Selenium WebDriver"""
        chrome_options = Options()
        chrome_options.add_argument('--headless') 
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        try:
            self.driver = webdriver.Chrome(options=chrome_options)
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        except Exception as e:
            logging.error(f"Erreur lors de l'initialisation de Selenium: {e}")
            self.use_selenium = False

    def is_masked(self, text):
        if not text:
            return True
        stars_ratio = text.count('*') / max(len(text), 1)
        return stars_ratio > 0.3

    def search_jobs_requests(self, keywords, location, max_pages=5):
        """Méthode avec requests (plus rapide mais plus limitée)"""
        jobs = []
        tpr_value = DURATION_MAP.get(self.duration, "r604800")
        
        for page in range(max_pages):
            base_url = "https://www.linkedin.com/jobs/search"
            params = {
                'keywords': keywords,   
                'location': location,
                'start': page * 25,
                'f_TPR': tpr_value,
            }
            
            url = f"{base_url}?" + "&".join([f"{k}={quote(str(v))}" for k, v in params.items()])
            
            try:
                print(f"Scraping page {page + 1}...")
                response = self.session.get(url, headers=self.headers, timeout=10)
                
                if response.status_code == 429:
                    print("Rate limit atteint, pause de 60 secondes...")
                    time.sleep(60)
                    continue
                
                if response.status_code != 200:
                    print(f"Erreur HTTP {response.status_code}")
                    continue
                
                soup = BeautifulSoup(response.content, 'html.parser')
                job_cards = soup.find_all('div', class_='base-card')
                
                if not job_cards:
                    print("Aucune offre trouvée sur cette page")
                    break
                
                for card in job_cards:
                    try:
                        job = self.extract_job_info_bs4(card)
                        if job:
                            jobs.append(job)
                    except Exception as e:
                        logging.error(f"Erreur lors de l'extraction d'une offre: {e}")
                        continue
                
                time.sleep(random.uniform(2, 5))
                
            except Exception as e:
                logging.error(f"Erreur lors du scraping de la page {page}: {e}")
                continue
        
        return jobs

    def search_jobs_selenium(self, keywords, location, max_pages=3):
        """Méthode avec Selenium (plus robuste mais plus lente)"""
        if not hasattr(self, 'driver'):
            print("Selenium n'est pas configuré")
            return []
        
        jobs = []
        tpr_value = DURATION_MAP.get(self.duration, "r604800")
        base_url = (
            f"https://www.linkedin.com/jobs/search"
            f"?keywords={quote(keywords)}&location={quote(location)}&f_TPR={tpr_value}"
        )
        
        try:
            self.driver.get(base_url)
            time.sleep(3)
            
            for page in range(max_pages):
                print(f"Scraping page {page + 1} avec Selenium...")
                
                wait = WebDriverWait(self.driver, 10)
                job_cards = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, '[data-entity-urn*="urn:li:fsu_jobPosting"]')))
                
                for card in job_cards:
                    try:
                        job = self.extract_job_info_selenium(card)
                        if job:
                            jobs.append(job)
                    except Exception as e:
                        logging.error(f"Erreur lors de l'extraction: {e}")
                        continue
                
                try:
                    next_button = self.driver.find_element(By.CSS_SELECTOR, 'button[aria-label="Page suivante"]')
                    if next_button.is_enabled():
                        next_button.click()
                        time.sleep(3)
                    else:
                        break
                except:
                    break
                    
        except Exception as e:
            logging.error(f"Erreur Selenium: {e}")
        
        return jobs
    
    def extract_job_info_bs4(self, card):
        """Extraire les informations d'une offre avec BeautifulSoup"""
        try:
            title_elem = card.find('h3', class_='base-search-card__title')
            title = title_elem.get_text().strip() if title_elem else "N/A"
            
            company_elem = card.find('h4', class_='base-search-card__subtitle')
            company = company_elem.get_text().strip() if company_elem else "N/A"
            
            location_elem = card.find('span', class_='job-search-card__location')
            location = location_elem.get_text().strip() if location_elem else "N/A"
            
            if self.is_masked(title) or self.is_masked(company) or self.is_masked(location):
                return None
            
            link_elem = card.find('a', class_='base-card__full-link')
            link = link_elem['href'] if link_elem else "N/A"
            
            date_elem = card.find('time')
            date_raw = date_elem['datetime'] if date_elem else "N/A"
            date_text = date_elem.get_text().strip() if date_elem else "N/A"

            try:
                dt = datetime.fromisoformat(date_raw.replace("Z", "+00:00"))
                date_formatted = dt.strftime("%d/%m/%Y")
            except:
                date_formatted = date_raw

            return {
                'titre': title,
                'entreprise': company,
                'localisation': location,
                'lien': link,
                'date_publication': date_formatted,
                'delai_publication': date_text,
            }
        except Exception as e:
            logging.error(f"Erreur extraction BeautifulSoup: {e}")
            return None

    def extract_job_info_selenium(self, card):
        """Extraire les informations d'une offre avec Selenium"""
        try:
            title = card.find_element(By.CSS_SELECTOR, 'h3').text.strip()
            company = card.find_element(By.CSS_SELECTOR, 'h4').text.strip()
            location = card.find_element(By.CSS_SELECTOR, '.job-search-card__location').text.strip()
            link = card.find_element(By.TAG_NAME, 'a').get_attribute('href')
            
            try:
                date = card.find_element(By.TAG_NAME, 'time').get_attribute('datetime')
            except:
                date = "N/A"
            
            return {
                'titre': title,
                'entreprise': company,
                'localisation': location,
                'lien': link,
                'date_publication': date,
            }
        except Exception as e:
            logging.error(f"Erreur extraction Selenium: {e}")
            return None

    def save_to_excel_with_colors(self, jobs, filename="linkedin_jobs.xlsx"):
        """Sauvegarder en Excel avec couleurs et formatage"""
        if not OPENPYXL_AVAILABLE:
            print("ERREUR: openpyxl non disponible - impossible de générer le fichier Excel")
            return
        
        if not jobs:
            print("Aucune offre à sauvegarder")
            return
        
        df = pd.DataFrame(jobs)
        
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].astype(str).str.strip()
                df[col] = df[col].str.replace('\n', ' ', regex=False)
                df[col] = df[col].str.replace('\r', ' ', regex=False)
                df[col] = df[col].str.replace('\t', ' ', regex=False)
                df[col] = df[col].replace('N/A', '')
        
        column_order = ['location_recherche', 'titre', 'entreprise', 'localisation', 'date_publication', 'delai_publication', 'lien']
        df = df.reindex(columns=[col for col in column_order if col in df.columns])
        
        df.rename(columns={
            'location_recherche': 'Ville recherchée',
            'titre': 'Titre du poste',
            'entreprise': 'Entreprise',
            'localisation': 'Localisation',
            'date_publication': 'Date',
            'delai_publication': 'Publié',
            'lien': 'Lien LinkedIn',
        }, inplace=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Offres LinkedIn"
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        light_blue = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
        light_green = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
        recent_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        old_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    if c_idx == 6 and isinstance(value, str):
                        if 'minute' in value.lower() or ('heure' in value.lower() and any(h in value for h in ['1 heure', '2 heure', '3 heure'])):
                            cell.fill = recent_fill
                        elif 'jour' in value.lower() or 'semaine' in value.lower():
                            cell.fill = old_fill
                    
                    if c_idx == 7 and value and value != '':
                        original_link = jobs[r_idx-2].get('lien', '') if r_idx-2 < len(jobs) else ''
                        if original_link and original_link != 'N/A':
                            cell.value = "*LIEN*"
                            cell.hyperlink = original_link
                            cell.font = Font(color="0000FF", underline="single")
                    
                    if r_idx % 2 == 0:
                        if not cell.fill.start_color.rgb or cell.fill.start_color.rgb == '00000000':
                            cell.fill = light_blue
                    else:
                        if not cell.fill.start_color.rgb or cell.fill.start_color.rgb == '00000000':
                            cell.fill = light_green
        
        column_widths = {
            'A': 20,
            'B': 40,
            'C': 25,
            'D': 20,
            'E': 12,
            'F': 15,
            'G': 12,
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        wb.save(filename)
        print(f"\n✓ Sauvegarde de {len(jobs)} offres dans {filename}")
        print("Fichier Excel avec couleurs créé :")
        print("  • En-têtes bleus avec texte blanc")
        print("  • Lignes alternées bleu/vert clair")
        print("  • Offres récentes en vert, anciennes en rouge")
        print("  • Liens cliquables '*LIEN*'")

    def close(self):
        """Fermer le navigateur Selenium si utilisé"""
        if hasattr(self, 'driver'):
            self.driver.quit()

    def deduplicate_jobs(self, jobs):
        """Supprimer les doublons par titre + entreprise + localisation avec logs"""
        seen = set()
        unique_jobs = []
        
        for job in jobs:
            key = (job.get("titre"), job.get("entreprise"), job.get("localisation"))
            if key not in seen:
                seen.add(key)
                unique_jobs.append(job)
            else:
                print(f"[DUPLICATE] supprimé: {job.get('titre')} @ {job.get('entreprise')} ({job.get('localisation')})")
        
        return unique_jobs

    def parse_delay_to_minutes(self, delay_text):
        """Convertir le délai français en minutes pour le tri"""
        if not delay_text or delay_text == "N/A":
            return float('inf')
        
        delay_text = delay_text.lower().strip()
        
        number_match = re.search(r'(\d+)', delay_text)
        if not number_match:
            return float('inf')
        
        number = int(number_match.group(1))
        
        if 'minute' in delay_text:
            return number
        elif 'heure' in delay_text:
            return number * 60
        elif 'jour' in delay_text:
            return number * 60 * 24
        elif 'semaine' in delay_text:
            return number * 60 * 24 * 7
        elif 'mois' in delay_text:
            return number * 60 * 24 * 30
        else:
            return float('inf')


def parse_args():
    parser = argparse.ArgumentParser(
        description="LinkedIn Job Scraper",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemples :
  python linkedin_scraper.py                  # 1 semaine (défaut)
  python linkedin_scraper.py --duration 1d    # 24 dernières heures
  python linkedin_scraper.py --duration 1w    # 7 derniers jours
        """
    )
    parser.add_argument(
        "--duration",
        choices=["1d", "1w"],
        default="1w",
        help="Durée de recherche des offres : 1d = 24h, 1w = 7 jours (défaut: 1w)",
    )
    parser.add_argument(
        "--pages",
        type=int,
        default=2,
        help="Nombre de pages à scraper par recherche (défaut: 2)",
    )
    parser.add_argument(
        "--selenium",
        action="store_true",
        help="Utiliser Selenium au lieu de requests",
    )
    parser.add_argument(
        "--output",
        default="linkedin_multiple_jobs.xlsx",
        help="Nom du fichier Excel de sortie (défaut: linkedin_multiple_jobs.xlsx)",
    )
    return parser.parse_args()


def main():
    args = parse_args()

    duration_label = "24 dernières heures" if args.duration == "1d" else "7 derniers jours"
    print(f"⏱  Durée de recherche : {duration_label} (--duration {args.duration})")
    print(f"📄 Pages par recherche : {args.pages}")
    print(f"💾 Fichier de sortie   : {args.output}")

    logging.basicConfig(level=logging.INFO)
    scraper = LinkedInJobScraper(use_selenium=args.selenium, duration=args.duration)
    
    try:
        keywords_list = [
            # "full stack developer",
            # "web developer",
            # "software engineer",
            # "project chief",
            # "IT consultant",
            "frontend developer",
            "software engineer web",
            "Sales Engineer",
            "SDR / BDR"
            "IT Analyst"
        ]

        locations_list = [
            # HARD TO FIND JOBS LOCATIONS
            "Geneva, Switzerland",
            "Lausanne, Switzerland",
            # "Zurich, Switzerland",
            "Rabat, Morocco",
            "Casablanca, Morocco",
            # EASY TO FIND JOBS LOCATIONS
            # "Paris, France",
            # "Toulouse, France",
            # "Taiwan",
        ]
        
        all_jobs = []

        for location in locations_list:
            print(f"\n{'='*60}")
            print(f"LOCALISATION: {location}")
            print(f"{'='*60}")
            
            for keywords in keywords_list:
                print(f"\n  → Recherche: {keywords}")
                
                if scraper.use_selenium:
                    jobs = scraper.search_jobs_selenium(keywords, location, args.pages)
                else:
                    jobs = scraper.search_jobs_requests(keywords, location, args.pages)
                
                for job in jobs:
                    job['location_recherche'] = location
                
                all_jobs.extend(jobs)
                
                print(f"  Trouvé {len(jobs)} offres pour '{keywords}' à {location}")
                time.sleep(5)
        
        all_jobs = scraper.deduplicate_jobs(all_jobs)

        print(f"\n{'='*60}")
        print(f"TOTAL: {len(all_jobs)} offres trouvées")
        print(f"{'='*60}")

        if all_jobs:
            df = pd.DataFrame(all_jobs)
            df["delai_minutes"] = df["delai_publication"].apply(scraper.parse_delay_to_minutes)
            df = df.sort_values(by=["location_recherche", "delai_minutes"], ascending=[True, True])
            df = df.drop(columns=["delai_minutes"])

            all_jobs_sorted = df.to_dict(orient="records")
            scraper.save_to_excel_with_colors(all_jobs_sorted, args.output)
        else:
            print("Aucune offre trouvée")

    except KeyboardInterrupt:
        print("\nArrêt du scraping...")
    
    except Exception as e:
        print(f"Erreur: {e}")
    
    finally:
        scraper.close()


if __name__ == "__main__":
    main()